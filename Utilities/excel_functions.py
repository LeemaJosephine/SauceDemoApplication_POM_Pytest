from openpyxl import load_workbook

class ExcelFunction:

    def __init__(self, file_name, sheet_name):
        self.file = file_name
        self.sheet = sheet_name
        self.workbook = load_workbook(self.file)
        self.sheet = self.workbook[sheet_name]

    # Write data in excel
    def write_data(self , row_number, column_number, data):
        self.sheet.cell(row=row_number, column=column_number).value = data
        self.workbook.save(self.file)

    # Read data from excel
    def get_data_from_excel(self):
        rows = self.sheet.max_row
        columns = self.sheet.max_column
        data = []

        for row in range(2, rows + 1):  # Skipping header row
            row_data = []
            for col in range(1, columns + 1):
                value = self.sheet.cell(row=row, column=col).value
                if value is None:
                    value = ""  # Explicitly handling nulls as empty strings
                row_data.append(value)
            data.append(tuple(row_data))

        return data



